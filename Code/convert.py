"""Utilities for converting JSON labeled cases into spreadsheets."""

import ast
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config as cfg


# Most converter constants are centralized in config.py.
ANSWER_KEY_CONFIG_FIELD = cfg.ANSWER_KEY_CONFIG_FIELD
CASE_ID_COLUMN_CANDIDATES = cfg.CASE_ID_COLUMN_CANDIDATES
MODALITY_COLUMN_CANDIDATES = cfg.MODALITY_COLUMN_CANDIDATES
REPORT_LIKE_TERMS = cfg.REPORT_LIKE_TERMS
NONE_ALIASES = cfg.NONE_ALIASES
BLANK_ALIASES = cfg.BLANK_ALIASES
RAW_REPORT_COLUMNS_TO_DROP_AFTER_MERGE = cfg.RAW_REPORT_COLUMNS_TO_DROP_AFTER_MERGE
CONFIDENCE_THRESHOLD_SUFFIX = cfg.CONFIDENCE_THRESHOLD_SUFFIX
CONFIDENCE_VOTE_COUNT_SUFFIX = cfg.CONFIDENCE_VOTE_COUNT_SUFFIX
CONFIDENCE_TOTAL_VOTES_SUFFIX = cfg.CONFIDENCE_TOTAL_VOTES_SUFFIX
CONFIDENCE_FINAL_LABEL_SUFFIX = cfg.CONFIDENCE_FINAL_LABEL_SUFFIX
CONFIDENCE_VOTES_SUFFIX = cfg.CONFIDENCE_VOTES_SUFFIX
REVIEW_TARGET_COLUMNS = cfg.REVIEW_TARGET_COLUMNS
REVIEW_FLAG_COLUMNS = cfg.REVIEW_FLAG_COLUMNS

# Review highlighting is intentionally separated into a small helper file so
# answer-key comparison code stays focused on red mismatch highlighting.
try:
    from spreadsheet_debug_checks import (
        find_review_excel_rows_by_severity,
        find_strange_excel_rows_with_reasons,
    )
except Exception:  # pragma: no cover - keep convert.py usable without helper file.
    def find_strange_excel_rows_with_reasons(output_df: pd.DataFrame) -> dict[int, list[str]]:
        return {}

    def find_review_excel_rows_by_severity(output_df: pd.DataFrame) -> dict[str, dict[int, list[str]]]:
        return {"red": {}, "yellow": find_strange_excel_rows_with_reasons(output_df)}

# ---------------------------------------------------------------------------
# Compatibility helpers
# ---------------------------------------------------------------------------


def _df_map(df: pd.DataFrame, func) -> pd.DataFrame:
    """Use DataFrame.map when available, otherwise fall back to applymap."""
    if hasattr(df, "map"):
        return df.map(func)
    return df.applymap(func)


# ---------------------------------------------------------------------------
# Cleaning / normalization
# ---------------------------------------------------------------------------


def clean_cell_value(value: Any) -> str:
    """Convert list and string values to readable format without brackets or quotes."""
    if isinstance(value, list):
        # Join list items with commas, removing quotes.
        return ", ".join(str(item).strip('"\'') for item in value)
    if isinstance(value, str):
        # Remove surrounding quotes if present.
        return value.strip('"\'')
    return str(value) if value is not None else ""


def _normalize_column_name(name: Any) -> str:
    """Normalize column names so CT_GT, CT GT, and CT.GT match."""
    return re.sub(r"[^a-z0-9]", "", str(name).strip().lower())


def _find_column(df: pd.DataFrame, candidates: tuple[str, ...]) -> str | None:
    """Return the first matching column from a candidate list."""
    normalized_to_original = {_normalize_column_name(col): col for col in df.columns}
    for candidate in candidates:
        match = normalized_to_original.get(_normalize_column_name(candidate))
        if match is not None:
            return match
    return None


def _matching_columns(df: pd.DataFrame, candidates: tuple[str, ...]) -> list[str]:
    """Return all columns matching the candidate list, in candidate priority order."""
    normalized_to_original = {_normalize_column_name(col): col for col in df.columns}
    matches: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        match = normalized_to_original.get(_normalize_column_name(candidate))
        if match is not None and match not in seen:
            matches.append(match)
            seen.add(match)
    return matches


def _normalize_case_id(value: Any) -> str:
    """Normalize case IDs for matching output rows to answer-key rows."""
    return str(value).strip().strip('"\'')


def _split_label_string(value: str) -> list[str]:
    """Split a cell value into labels while preserving labels such as LMCA/RMCA if present."""
    text = value.strip()
    if not text:
        return []

    # Handle strings that look like Python/JSON lists, e.g. "['RMCA', 'RICA']".
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, (list, tuple, set)):
                return [str(item) for item in parsed]
        except (SyntaxError, ValueError):
            pass

    return re.split(r"[,;\n]+", text)


def _normalize_label_piece(piece: Any) -> str | None:
    label = str(piece).strip().strip('"\'[](){}').upper().replace(" ", "")

    if label in BLANK_ALIASES:
        return None

    # Treat common answer-key negative wording as NONE instead of as a literal
    # label called NEGATIVE. Without this, rows with GT="negative" compare as
    # {"NEGATIVE"} != {"NONE"} and get falsely highlighted.
    if label in NONE_ALIASES:
        return "NONE"

    return label


def _normalize_label_set(value: Any) -> set[str]:
    """Normalize label cells for fair comparison between model output and GT."""
    if value is None:
        return set()

    if isinstance(value, float) and pd.isna(value):
        return set()

    if isinstance(value, (list, tuple, set)):
        pieces = [str(item) for item in value]
    else:
        pieces = _split_label_string(str(value))

    labels: set[str] = set()
    for piece in pieces:
        label = _normalize_label_piece(piece)
        if label:
            labels.add(label)

    # If NONE is mixed with a real territory, ignore NONE so
    # "NONE, RMCA" compares like "RMCA".
    if "NONE" in labels and len(labels) > 1:
        labels.remove("NONE")

    return labels


# ---------------------------------------------------------------------------
# Column selection
# ---------------------------------------------------------------------------


def _column_looks_like_report(df: pd.DataFrame, column: str, sample_size: int = 25) -> bool:
    """Return True when a candidate prediction column appears to contain report text.

    This prevents a raw report column named CT/CTA/CTP from being used as the
    model-prediction column.
    """
    if column not in df.columns:
        return False

    values = [
        str(value)
        for value in df[column].dropna().head(sample_size).tolist()
        if str(value).strip()
    ]
    if not values:
        return False

    report_term_hits = 0
    long_values = 0
    for value in values:
        upper_value = value.upper()
        if any(term in upper_value for term in REPORT_LIKE_TERMS):
            report_term_hits += 1
        if len(value) > 120:
            long_values += 1

    # A single report-looking cell is enough when the values are very long,
    # because label cells are normally short: NONE, RMCA, LMCA, etc.
    return report_term_hits > 0 or long_values >= max(2, len(values) // 2)


def _find_prediction_column(output_df: pd.DataFrame, modality: str) -> str | None:
    """Find the best prediction column while avoiding raw report-text columns."""
    candidates = MODALITY_COLUMN_CANDIDATES[modality]["prediction"]
    matches = _matching_columns(output_df, candidates)

    for column in matches:
        if not _column_looks_like_report(output_df, column):
            return column

    # If every match looked like a report, return None instead of causing many
    # false-positive highlights.
    return None


# ---------------------------------------------------------------------------
# Answer-key comparison
# ---------------------------------------------------------------------------


def _read_answer_key(answer_key_path: Path) -> pd.DataFrame:
    """Read an Excel or CSV answer key into a DataFrame."""
    suffix = answer_key_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(answer_key_path)
    if suffix == ".csv":
        return pd.read_csv(answer_key_path)
    raise ValueError(f"Unsupported answer key format: {answer_key_path.suffix}")


def _configured_answer_key_path() -> Path | None:
    """Return config.GROUND_TRUTH_FILE only when it exists and is not blank."""
    try:
        import config  # type: ignore
    except Exception:
        return None

    raw_path = getattr(config, ANSWER_KEY_CONFIG_FIELD, None)

    # Guard: no configured answer key means conversion should still run,
    # just without grading/red answer-key highlighting.
    if raw_path is None or str(raw_path).strip() == "":
        return None

    answer_key_path = Path(raw_path)
    if not answer_key_path.exists():
        print(f"Answer key configured but not found; skipping grading: {answer_key_path}")
        return None

    return answer_key_path


def _resolve_answer_key_path(answer_key_path: str | Path | None = None) -> Path | None:
    """Resolve an optional answer key path.

    Returns None when no key is provided, config.GROUND_TRUTH_FILE is blank,
    or the key file does not exist. This keeps JSON -> Excel conversion from
    failing just because grading cannot run.
    """
    if answer_key_path is None:
        return _configured_answer_key_path()

    if str(answer_key_path).strip() == "":
        return None

    resolved_path = Path(answer_key_path)
    if not resolved_path.exists():
        print(f"Answer key not found; skipping grading: {resolved_path}")
        return None

    return resolved_path


def find_wrong_excel_cells(
    output_df: pd.DataFrame,
    answer_key_path: str | Path | None = None,
) -> dict[int, set[str]]:
    """Return Excel row numbers and output columns that should be highlighted red.

    If no answer key is configured or found, return an empty dict so the
    converter proceeds without grading.

    Only the case ID cell and wrong prediction cell(s) are returned.

    Example return value:
        {
            2: {"case_id", "CT_GT"},
            7: {"case_id", "CTA_GT", "Combined_GT"},
        }

    The row numbers are Excel row numbers, so DataFrame index 0 is row 2.
    """
    resolved_answer_key_path = _resolve_answer_key_path(answer_key_path)
    if resolved_answer_key_path is None:
        return {}

    try:
        answer_key_df = _read_answer_key(resolved_answer_key_path)
    except Exception as exc:
        print(f"Could not read answer key; skipping grading: {resolved_answer_key_path} ({exc})")
        return {}

    answer_key_df = _df_map(answer_key_df, clean_cell_value)

    output_case_col = _find_column(output_df, CASE_ID_COLUMN_CANDIDATES)
    answer_key_case_col = _find_column(answer_key_df, CASE_ID_COLUMN_CANDIDATES)

    if output_case_col is None or answer_key_case_col is None:
        print("Could not find a case ID column in the output or answer key, so cell highlighting was skipped.")
        return {}

    comparisons: list[tuple[str, str, str]] = []
    for modality, candidates in MODALITY_COLUMN_CANDIDATES.items():
        pred_col = _find_prediction_column(output_df, modality)
        gt_col = _find_column(answer_key_df, candidates["ground_truth"])
        if pred_col is not None and gt_col is not None:
            comparisons.append((modality, pred_col, gt_col))

    if not comparisons:
        print("No matching prediction/answer-key columns were found, so cell highlighting was skipped.")
        return {}

    answer_key_by_case = {
        _normalize_case_id(row[answer_key_case_col]): row
        for _, row in answer_key_df.iterrows()
        if _normalize_case_id(row[answer_key_case_col])
    }

    wrong_cells: dict[int, set[str]] = {}
    for df_index, output_row in output_df.iterrows():
        case_id = _normalize_case_id(output_row[output_case_col])
        answer_key_row = answer_key_by_case.get(case_id)
        if answer_key_row is None:
            continue

        excel_row_number = int(df_index) + 2  # +2 because Excel has a 1-row header.

        for _modality, pred_col, gt_col in comparisons:
            expected_labels = _normalize_label_set(answer_key_row[gt_col])

            # Blank GT cells should not cause a cell to be marked wrong.
            if not expected_labels:
                continue

            predicted_labels = _normalize_label_set(output_row[pred_col])
            if predicted_labels != expected_labels:
                wrong_cells.setdefault(excel_row_number, set()).update({output_case_col, pred_col})

    if wrong_cells:
        wrong_rows = len(wrong_cells)
        wrong_answer_cells = sum(max(0, len(cols) - 1) for cols in wrong_cells.values())
        print(
            f"Highlighted {wrong_rows} case_id cell(s) and "
            f"{wrong_answer_cells} wrong answer cell(s) red based on the answer key."
        )

    return wrong_cells


def find_wrong_excel_rows(
    output_df: pd.DataFrame,
    answer_key_path: str | Path | None = None,
) -> set[int]:
    """Backward-compatible wrapper returning rows with at least one wrong answer."""
    return set(find_wrong_excel_cells(output_df, answer_key_path).keys())



# ---------------------------------------------------------------------------
# Report/reasoning display formatting
# ---------------------------------------------------------------------------

RAW_REPORT_COLUMNS_TO_DROP_AFTER_MERGE = (
    "CT_Report",
    "CTA_Report",
    "CTP_Report",
    "MRI_Report",
)

# Confidence/review columns that are useful in JSON but noisy in the final
# reviewer spreadsheet.  Thresholds are constant for a run, vote counts are
# merged into a compact n/total column, and *_final_label duplicates the label
# columns that are already shown near the front of the sheet.
CONFIDENCE_THRESHOLD_SUFFIX = "_confidence_threshold"
CONFIDENCE_VOTE_COUNT_SUFFIX = "_confidence_vote_count"
CONFIDENCE_TOTAL_VOTES_SUFFIX = "_confidence_total_votes"
CONFIDENCE_FINAL_LABEL_SUFFIX = "_final_label"
CONFIDENCE_VOTES_SUFFIX = "_confidence_votes"

# Original JSON label fields mapped to the display spreadsheet answer and
# report/reasoning columns.  These targets are used for cell-level review
# highlighting, especially low-confidence flags.
REVIEW_TARGET_COLUMNS = {
    "CT_Original_GT": ("CT_Original_GT", "CT Original Report/Reasoning"),
    "CT_GT": ("CT_GT", "CT Report/Reasoning"),
    "CTA_GT": ("CTA_GT", "CTA Report/Reasoning"),
    "CTP_GT": ("CTP_GT", "CTP Report/Reasoning"),
    "Combined_GT": ("Combined_GT", "Combined Report/Reasoning"),
}

REVIEW_FLAG_COLUMNS = (
    "Review_Flags_Red",
    "Review_Flags_Yellow",
    "Review_Flags",
)


def _clean_display_text(value: Any) -> str:
    """Return display text while treating common empty spreadsheet values as blank."""
    if value is None:
        return ""
    try:
        if isinstance(value, float) and pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text in {"", "nan", "NaN", "None", "NONE", "null", "NULL"}:
        return ""
    return text


def _first_nonempty(row: pd.Series, columns: tuple[str, ...]) -> str:
    for column in columns:
        if column in row.index:
            text = _clean_display_text(row.get(column, ""))
            if text:
                return text
    return ""


def _format_vote_piece(value: Any) -> str:
    """Return vote-count display text without trailing .0 when possible."""
    text = _clean_display_text(value)
    if not text:
        return ""
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except Exception:
        pass
    return text


def _format_confidence_vote_pair(vote_count: Any, total_votes: Any) -> str:
    """Return a compact confidence vote string such as 5/7 or 7/7."""
    vote_text = _format_vote_piece(vote_count)
    total_text = _format_vote_piece(total_votes)
    if vote_text and total_text:
        return f"{vote_text}/{total_text}"
    return vote_text or total_text


def _format_confidence_columns_for_display(output_df: pd.DataFrame) -> pd.DataFrame:
    """Return a display DataFrame with confidence columns made reviewer-friendly.

    Spreadsheet-only cleanup:
    - Drop *_confidence_threshold because the value is constant for the run.
    - Merge *_confidence_vote_count and *_confidence_total_votes into
      *_confidence_votes with values like 5/7.
    - Drop *_final_label because the final answer labels are already shown in
      the main *_GT columns near the front of the sheet.

    The original cleaned DataFrame is still used for grading/review logic, so
    this only affects the final spreadsheet view.
    """
    display_df = output_df.copy()
    columns_to_drop: set[str] = set()

    for vote_col in list(display_df.columns):
        if not str(vote_col).endswith(CONFIDENCE_VOTE_COUNT_SUFFIX):
            continue

        prefix = str(vote_col)[: -len(CONFIDENCE_VOTE_COUNT_SUFFIX)]
        total_col = f"{prefix}{CONFIDENCE_TOTAL_VOTES_SUFFIX}"
        votes_col = f"{prefix}{CONFIDENCE_VOTES_SUFFIX}"

        if total_col not in display_df.columns:
            continue

        votes_values = [
            _format_confidence_vote_pair(vote_count, total_votes)
            for vote_count, total_votes in zip(display_df[vote_col], display_df[total_col])
        ]

        insert_at = display_df.columns.get_loc(vote_col)
        if votes_col in display_df.columns:
            display_df[votes_col] = votes_values
        else:
            display_df.insert(insert_at, votes_col, votes_values)

        columns_to_drop.update({vote_col, total_col})

    for column in list(display_df.columns):
        column_text = str(column)
        if column_text.endswith(CONFIDENCE_THRESHOLD_SUFFIX) or column_text.endswith(CONFIDENCE_FINAL_LABEL_SUFFIX):
            columns_to_drop.add(column)

    drop_columns = [column for column in display_df.columns if column in columns_to_drop]
    if drop_columns:
        display_df = display_df.drop(columns=drop_columns)

    return display_df


def _report_block(row: pd.Series, sections: tuple[tuple[str, tuple[str, ...]], ...]) -> str:
    parts: list[str] = []
    for title, columns in sections:
        text = _first_nonempty(row, columns)
        if text:
            parts.append(f"{title}:\n{text}")
    return "\n\n".join(parts)


def _combine_report_and_reasoning(report_text: str, reasoning_text: str) -> str:
    report_text = _clean_display_text(report_text)
    reasoning_text = _clean_display_text(reasoning_text)
    if report_text and reasoning_text:
        return f"{report_text}\n\n{reasoning_text}"
    return report_text or reasoning_text


def _format_report_reasoning_columns(output_df: pd.DataFrame) -> pd.DataFrame:
    """Return an Excel-display DataFrame with report text merged into reasoning cells.

    Each *_reasoning column becomes a readable Report/Reasoning column.  The
    cell value is the relevant report text, then two blank-line-separated
    newlines, then the model reasoning.  This keeps the evidence and explanation
    together when reviewing the spreadsheet.
    """
    display_df = output_df.copy()

    builders = {
        "CT_Original_GT_reasoning": (
            "CT Original Report/Reasoning",
            lambda row: _first_nonempty(row, ("CT_Report", "New_CT_Report")),
        ),
        "CT_GT_reasoning": (
            "CT Report/Reasoning",
            lambda row: _first_nonempty(row, ("New_CT_Report", "CT_Report")),
        ),
        "CT_Sanitization_reasoning": (
            "CT Sanitization Report/Reasoning",
            lambda row: _report_block(row, (
                ("Original CT Report", ("CT_Report",)),
                ("Sanitized CT Report", ("New_CT_Report",)),
            )),
        ),
        "CTA_GT_reasoning": (
            "CTA Report/Reasoning",
            lambda row: _first_nonempty(row, ("CTA_Report",)),
        ),
        "CTP_GT_reasoning": (
            "CTP Report/Reasoning",
            lambda row: _first_nonempty(row, ("CTP_Report",)),
        ),
        "CT_Combined_GT_reasoning": (
            "Combined Report/Reasoning",
            lambda row: _report_block(row, (
                ("CT Report", ("New_CT_Report", "CT_Report")),
                ("CTA Report", ("CTA_Report",)),
                ("CTP Report", ("CTP_Report",)),
                ("MRI Report", ("MRI_Report",)),
            )),
        ),
    }

    for source_column, (display_column, report_builder) in builders.items():
        if source_column not in display_df.columns:
            continue

        display_values = []
        for _, row in display_df.iterrows():
            report_text = report_builder(row)
            reasoning_text = row.get(source_column, "")
            display_values.append(_combine_report_and_reasoning(report_text, reasoning_text))

        # Preserve the original column position by replacing the reasoning column
        # in place, then renaming it to the friendlier display header.
        display_df[source_column] = display_values
        display_df = display_df.rename(columns={source_column: display_column})

    drop_columns = [
        column
        for column in RAW_REPORT_COLUMNS_TO_DROP_AFTER_MERGE
        if column in display_df.columns
    ]
    if drop_columns:
        display_df = display_df.drop(columns=drop_columns)

    return display_df


def _format_current_prompt_columns_for_display(output_df: pd.DataFrame) -> pd.DataFrame:
    """Append active prompt/instruction columns to the Excel display DataFrame.

    The DSPy version does not always store a literal prompt in each JSON case.
    Instead, the active task instructions live in config.py as
    CT_SIGNATURE_INSTRUCTIONS, CTA_SIGNATURE_INSTRUCTIONS, CTP_SIGNATURE_INSTRUCTIONS,
    and COMBINED_SIGNATURE_INSTRUCTIONS.  This function copies those active
    config prompts into the final spreadsheet so every output file records the
    prompt/instruction context used for the run.
    """
    display_df = output_df.copy()

    if not bool(getattr(cfg, "INCLUDE_CURRENT_PROMPT_COLUMNS", False)):
        return display_df

    prompt_columns = getattr(cfg, "CURRENT_PROMPT_COLUMNS", {})
    if not isinstance(prompt_columns, dict) or not prompt_columns:
        return display_df

    for column_name, prompt_text in prompt_columns.items():
        column_name = str(column_name).strip()
        if not column_name:
            continue

        # Clean prompt text for spreadsheet display but do not treat "NONE" as
        # blank here because a prompt may legitimately mention the NONE label.
        prompt = "" if prompt_text is None else str(prompt_text).strip()
        if column_name in display_df.columns:
            display_df[column_name] = prompt
        else:
            display_df[column_name] = prompt

    return display_df


# ---------------------------------------------------------------------------
# Cell-level review highlighting helpers
# ---------------------------------------------------------------------------


def _truthy_cell(value: Any) -> bool:
    """Return True for common spreadsheet truthy values."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if pd.isna(value):
                return False
        except Exception:
            pass
        return value != 0
    text = _clean_display_text(value).lower()
    return text in {"true", "t", "yes", "y", "1"}


def _false_confidence_cell(value: Any) -> bool:
    """Return True only when a confidence field is explicitly false."""
    if isinstance(value, bool):
        return not value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            if pd.isna(value):
                return False
        except Exception:
            pass
        return value == 0
    text = _clean_display_text(value).lower()
    return text in {"false", "f", "no", "n", "0"}


def _add_display_target_columns(
    output: set[str],
    display_df: pd.DataFrame,
    *,
    case_column: str | None,
    label_field: str | None = None,
    include_reasoning: bool = True,
) -> None:
    """Add case ID + answer + reasoning display columns when they exist."""
    if case_column and case_column in display_df.columns:
        output.add(case_column)

    if not label_field:
        return

    answer_column, reasoning_column = REVIEW_TARGET_COLUMNS.get(label_field, (label_field, ""))
    if answer_column in display_df.columns:
        output.add(answer_column)
    if include_reasoning and reasoning_column in display_df.columns:
        output.add(reasoning_column)


def _review_label_fields_from_reason(reason: str) -> list[str]:
    """Infer label fields to highlight from a deterministic review reason."""
    text = str(reason or "").lower()
    fields: list[str] = []

    # Check more specific CT-original wording before generic CT wording.
    if any(term in text for term in ("ct original", "ct_original_gt")):
        fields.append("CT_Original_GT")
    if any(term in text for term in ("combined", "combined_gt", "combined gt")):
        fields.append("Combined_GT")
    if any(term in text for term in ("cta", "cta_gt", "cta gt")):
        fields.append("CTA_GT")
    if any(term in text for term in ("ctp", "ctp_gt", "ctp gt", "tmax", "perfusion")):
        fields.append("CTP_GT")
    if any(term in text for term in (
        "ct_gt",
        "ct gt",
        "ct label",
        "ct report",
        "new_ct_report",
        "sanitization",
        "sanitized",
    )):
        fields.append("CT_GT")

    # Reasons that mention CT but not CT_GT or CT original usually refer to CT.
    if not fields and re.search(r"\bct\b", text):
        fields.append("CT_GT")

    deduped: list[str] = []
    for field in fields:
        if field not in deduped:
            deduped.append(field)
    return deduped


def _find_low_confidence_excel_cells(
    output_df: pd.DataFrame,
    display_df: pd.DataFrame,
) -> dict[int, set[str]]:
    """Return targeted cells for any *_is_confident == False field.

    Example: Combined_GT_is_confident = False highlights case_id,
    Combined_GT, and Combined Report/Reasoning instead of the whole row.
    """
    case_column = _find_column(display_df, CASE_ID_COLUMN_CANDIDATES)
    targeted_cells: dict[int, set[str]] = {}

    for df_index, row in output_df.iterrows():
        excel_row_number = int(df_index) + 2
        for label_field in REVIEW_TARGET_COLUMNS:
            confidence_column = f"{label_field}_is_confident"
            if confidence_column not in output_df.columns:
                continue
            if not _false_confidence_cell(row.get(confidence_column, "")):
                continue

            row_targets = targeted_cells.setdefault(excel_row_number, set())
            _add_display_target_columns(
                row_targets,
                display_df,
                case_column=case_column,
                label_field=label_field,
                include_reasoning=True,
            )

    return targeted_cells


def _find_review_excel_cells(
    output_df: pd.DataFrame,
    display_df: pd.DataFrame,
    review_reasons: dict[int, list[str]],
) -> dict[int, set[str]]:
    """Return targeted cells for deterministic review reasons.

    Low-confidence targets are handled exactly from *_is_confident columns.
    Other review reasons are mapped to the most likely label/reasoning columns;
    if a reason cannot be mapped, only case_id and Review_Flags columns are
    highlighted instead of painting the entire row.
    """
    case_column = _find_column(display_df, CASE_ID_COLUMN_CANDIDATES)
    targeted_cells = _find_low_confidence_excel_cells(output_df, display_df)

    for excel_row_number, reasons in review_reasons.items():
        row_targets = targeted_cells.setdefault(excel_row_number, set())
        if case_column and case_column in display_df.columns:
            row_targets.add(case_column)

        mapped_any_field = False
        for reason in reasons:
            for label_field in _review_label_fields_from_reason(reason):
                mapped_any_field = True
                _add_display_target_columns(
                    row_targets,
                    display_df,
                    case_column=case_column,
                    label_field=label_field,
                    include_reasoning=True,
                )

        # Fallback for general review flags that are not tied to a single modality.
        if not mapped_any_field:
            for column in REVIEW_FLAG_COLUMNS:
                if column in display_df.columns:
                    row_targets.add(column)

        if not row_targets:
            targeted_cells.pop(excel_row_number, None)

    return targeted_cells


# ---------------------------------------------------------------------------
# Debug summary sheet
# ---------------------------------------------------------------------------


def _write_debug_summary_sheet(
    excel_path: str,
    output_df: pd.DataFrame,
    wrong_cells: dict[int, set[str]],
    strange_reasons: dict[int, list[str]],
    red_review_reasons: dict[int, list[str]] | None = None,
    sheet_name: str = "Cases_For_Review",
) -> None:
    """Create a compact sheet showing one row per case that needs attention.

    A case may have multiple red answer mismatches and multiple yellow review
    reasons.  The Cases_For_Review sheet intentionally writes that case only
    once and combines the fields/reasons into one readable row.

    Reviewer-friendly layout:
    - Current_Value is placed where Issue_Type used to be so the reviewer sees
      the current model value immediately.
    - Field contains only the report type that needs review, such as CT, CTA,
      CTP, or CT Combined, instead of raw JSON column names.
    - Issue_Type is still preserved later in the row for filtering.
    """
    from openpyxl import load_workbook

    red_review_reasons = red_review_reasons or {}

    if not wrong_cells and not strange_reasons and not red_review_reasons:
        return

    wb = load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    case_col = _find_column(output_df, CASE_ID_COLUMN_CANDIDATES)
    headers = ["case_id", "Excel_Row", "Current_Value", "Field", "Issue_Type", "Reason"]
    ws.append(headers)

    FIELD_DISPLAY_NAMES = {
        "CT_Original_GT": "CT",
        "CT_GT": "CT",
        "CT_Sanitization_reasoning": "CT",
        "CTA_GT": "CTA",
        "CTP_GT": "CTP",
        "Combined_GT": "CT Combined",
        "CT_Combined_GT": "CT Combined",
    }

    def row_value(excel_row_number: int, column: str) -> str:
        df_index = excel_row_number - 2
        if df_index < 0 or df_index >= len(output_df) or column not in output_df.columns:
            return ""
        return clean_cell_value(output_df.iloc[df_index].get(column, ""))

    def case_id_for_row(excel_row_number: int) -> str:
        if case_col is None:
            return ""
        return row_value(excel_row_number, case_col)

    def dedupe_text(items: list[str]) -> list[str]:
        seen: set[str] = set()
        deduped: list[str] = []
        for item in items:
            text = str(item).strip()
            if text and text not in seen:
                deduped.append(text)
                seen.add(text)
        return deduped

    def review_field_for_column(column_name: str) -> str:
        """Convert raw output columns into reviewer-friendly report types."""
        if column_name in FIELD_DISPLAY_NAMES:
            return FIELD_DISPLAY_NAMES[column_name]

        normalized = _normalize_column_name(column_name)
        if "combined" in normalized:
            return "CT Combined"
        if normalized.startswith("cta"):
            return "CTA"
        if normalized.startswith("ctp"):
            return "CTP"
        if normalized.startswith("ct") or "sanitization" in normalized:
            return "CT"
        return "Review"

    def review_fields_from_reason(reason: str) -> list[str]:
        """Infer the report type(s) from deterministic review/check text."""
        text = str(reason).lower()
        fields: list[str] = []

        # Check Combined before CT so Combined_GT does not become a generic CT review.
        if any(term in text for term in ("combined_gt", "combined gt", "combined")):
            fields.append("CT Combined")
        if any(term in text for term in (
            "ct_original_gt",
            "ct original",
            "ct_gt",
            "ct gt",
            "ct label",
            "ct report",
            "new_ct_report",
            "sanitization",
            "sanitized",
        )):
            fields.append("CT")
        if any(term in text for term in ("cta_gt", "cta gt", "cta report", "cta")):
            fields.append("CTA")
        if any(term in text for term in ("ctp_gt", "ctp gt", "ctp report", "ctp", "tmax", "perfusion")):
            fields.append("CTP")

        return dedupe_text(fields) or ["Review"]

    red_priority_rows = set(wrong_cells) | set(red_review_reasons)
    all_rows = sorted(red_priority_rows | set(strange_reasons))
    for excel_row_number in all_rows:
        issue_types: list[str] = []
        fields: list[str] = []
        current_values: list[str] = []
        reasons: list[str] = []

        wrong_columns = [
            column_name
            for column_name in sorted(wrong_cells.get(excel_row_number, set()))
            if column_name != case_col
        ]
        if wrong_columns:
            issue_types.append("Answer mismatch")
            for column_name in wrong_columns:
                field_name = review_field_for_column(column_name)
                fields.append(field_name)
                current_values.append(
                    f"{field_name}: {row_value(excel_row_number, column_name)}"
                )
            reasons.append(
                "Prediction differs from the answer key; red cells show the exact field(s)."
            )

        high_priority_reasons = red_review_reasons.get(excel_row_number, [])
        review_reasons = strange_reasons.get(excel_row_number, [])
        if high_priority_reasons:
            issue_types.append("High-priority review")
            for reason in high_priority_reasons:
                fields.extend(review_fields_from_reason(reason))
            reasons.extend(high_priority_reasons)

        if review_reasons:
            issue_types.append("Review")
            for reason in review_reasons:
                fields.extend(review_fields_from_reason(reason))
            reasons.extend(review_reasons)

        if high_priority_reasons or review_reasons:
            review_flags = row_value(excel_row_number, "Review_Flags")
            if review_flags:
                current_values.append(f"Review_Flags: {review_flags}")

        ws.append([
            case_id_for_row(excel_row_number),
            excel_row_number,
            "\n".join(dedupe_text(current_values)),
            " | ".join(dedupe_text(fields)),
            " + ".join(dedupe_text(issue_types)),
            "\n".join(f"- {reason}" for reason in dedupe_text(reasons)),
        ])

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    strange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        excel_row_number = ws.cell(row=row_index, column=2).value
        try:
            excel_row_number = int(excel_row_number)
        except Exception:
            excel_row_number = -1
        fill = wrong_fill if excel_row_number in red_priority_rows else strange_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = wrap_alignment
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 28, "B": 10, "C": 38, "D": 18, "E": 20, "F": 85}
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45

    wb.save(excel_path)

# ---------------------------------------------------------------------------
# Excel output styling
# ---------------------------------------------------------------------------


def style_excel_sheet(
    excel_path: str,
    sheet_name: str = "Generated_Output",
    wrong_cells: dict[int, set[str]] | None = None,
    wrong_rows: set[int] | None = None,
    strange_rows: set[int] | None = None,
    red_review_rows: set[int] | None = None,
    red_review_cells: dict[int, set[str]] | None = None,
    yellow_review_cells: dict[int, set[str]] | None = None,
) -> None:
    """Apply styling and formatting to the Excel spreadsheet.

    wrong_cells highlights selected answer-key mismatch cells red.
    red_review_cells highlights targeted high-priority review cells red.
    yellow_review_cells highlights targeted warning/review cells yellow.
    When targeted review cells are provided, review flags no longer paint the
    entire row; they highlight case_id + the affected answer/reasoning cells.
    wrong_rows/red_review_rows/strange_rows are kept for backward compatibility.
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    wrong_cells = wrong_cells or {}
    wrong_rows = wrong_rows or set()
    strange_rows = strange_rows or set()
    red_review_rows = red_review_rows or set()
    red_review_cells = red_review_cells or {}
    yellow_review_cells = yellow_review_cells or {}

    header_to_col = {
        str(ws.cell(row=1, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }
    report_reasoning_col_indexes = {
        col_idx
        for header, col_idx in header_to_col.items()
        if "Report/Reasoning" in header or "Current Prompt" in header
    }

    def column_name_map_to_indexes(cells_by_row: dict[int, set[str]]) -> dict[int, set[int]]:
        indexes: dict[int, set[int]] = {}
        for row_idx, column_names in cells_by_row.items():
            for column_name in column_names:
                col_idx = header_to_col.get(str(column_name))
                if col_idx is not None:
                    indexes.setdefault(row_idx, set()).add(col_idx)
        return indexes

    wrong_cell_indexes = column_name_map_to_indexes(wrong_cells)
    red_review_cell_indexes = column_name_map_to_indexes(red_review_cells)
    yellow_review_cell_indexes = column_name_map_to_indexes(yellow_review_cells)
    targeted_review_mode = bool(red_review_cell_indexes or yellow_review_cell_indexes)

    # Define styles.
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alternating_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    strange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Apply header styling.
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Apply data styling and set row heights.
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        default_fill = alternating_fill if (row_idx - 2) % 2 == 0 else PatternFill()

        # Old behavior painted whole rows for review flags.  New behavior uses
        # targeted review cells, so low-confidence flags highlight case_id +
        # affected answer/reasoning only.  Keep row-level review coloring only
        # for older callers that do not pass red/yellow_review_cells.
        if not targeted_review_mode:
            if row_idx in red_review_rows:
                default_fill = wrong_fill
            elif row_idx in strange_rows:
                default_fill = strange_fill

        wrong_cols = wrong_cell_indexes.get(row_idx, set())
        red_review_cols = red_review_cell_indexes.get(row_idx, set())
        yellow_review_cols = yellow_review_cell_indexes.get(row_idx, set())

        for cell in row:
            if cell.column in wrong_cols or cell.column in red_review_cols:
                cell.fill = wrong_fill
            elif cell.column in yellow_review_cols:
                cell.fill = strange_fill
            elif not targeted_review_mode and not wrong_cells and row_idx in wrong_rows:
                # Backward compatibility for old callers using wrong_rows.
                cell.fill = wrong_fill
            else:
                cell.fill = default_fill

            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

        ws.row_dimensions[row_idx].height = 95 if report_reasoning_col_indexes else 30

    # Set header row height.
    ws.row_dimensions[1].height = 25

    # Auto-adjust column widths.
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        # Set width with some padding. Report/Reasoning and Current Prompt columns are intentionally
        # wider because they contain long report/reasoning or prompt text.
        if col_idx in report_reasoning_col_indexes:
            adjusted_width = 90
        else:
            adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)


# ---------------------------------------------------------------------------
# Public converter
# ---------------------------------------------------------------------------


def convert(
    json_path: str = "Files/labeled_cases.json",
    out_path: str | None = None,
    *,
    fmt: str = "xlsx",
    sheet_name: str = "Sheet1",
    answer_key_path: str | Path | None = None,
    highlight_wrong_rows: bool = True,
    highlight_strange_rows: bool = True,
    add_debug_sheet: bool = True,
) -> Path:
    """Load JSON from `json_path` and write it to a spreadsheet.

    Args:
        json_path: Path to the input JSON file (list of dicts or dict).
        out_path: Optional output file path. If omitted, uses the input
            filename with an appropriate extension in the same folder.
        fmt: Output format: 'xlsx' (default) or 'csv'.
        sheet_name: Excel sheet name when writing .xlsx.
        answer_key_path: Optional explicit answer-key path. When omitted,
            config.GROUND_TRUTH_FILE is used if it is present and not blank.
        highlight_wrong_rows: If True, mark mismatches red. Despite the legacy
            argument name, this now highlights only the case_id cell and wrong
            answer cell(s), not the entire row. Only applies to xlsx output.
        highlight_strange_rows: If True, highlight rows yellow when the output
            contains review-worthy / strange signals such as Needs_Review,
            Review_Flags, invalid labels, or CT labels changing after sanitization.
        add_debug_sheet: If True, add a Debug_Summary sheet listing red
            answer-key mismatches and yellow strange/review rows.

    Returns:
        Path to the written output file.
    """
    src = Path(json_path)
    # Failsafe.
    if not src.exists():
        raise FileNotFoundError(f"Input JSON not found: {src}")

    with src.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        try:
            df = pd.DataFrame.from_dict(data)
        except Exception:
            df = pd.json_normalize(data)
    else:
        df = pd.DataFrame(data)

    # Clean up cell values (remove brackets, quotes, etc.).
    df = _df_map(df, clean_cell_value)

    # Keep the original cleaned DataFrame for grading/debug logic because those
    # helpers expect the original *_reasoning field names.  Use a separate
    # display DataFrame for the spreadsheet so report text and reasoning appear
    # together in reviewer-friendly columns.  The spreadsheet view also gets
    # confidence-QOL cleanup: drop constant threshold/final-label columns and
    # merge vote_count + total_votes into one n/total column.
    display_base_df = _format_confidence_columns_for_display(df)
    display_df = _format_report_reasoning_columns(display_base_df)
    display_df = _format_current_prompt_columns_for_display(display_df)

    if out_path is None:
        out = src.with_suffix(".xlsx") if fmt == "xlsx" else src.with_suffix(".csv")
    else:
        out = Path(out_path)

    if fmt == "xlsx":
        wrong_cells = find_wrong_excel_cells(df, answer_key_path) if highlight_wrong_rows else {}
        if highlight_strange_rows:
            review_buckets = find_review_excel_rows_by_severity(df)
            red_review_reasons = review_buckets.get("red", {})
            yellow_review_reasons = review_buckets.get("yellow", {})
        else:
            red_review_reasons = {}
            yellow_review_reasons = {}

        # Red priority: if the same row has red and yellow review reasons, style
        # it red but still list both types in Cases_For_Review.  Main-sheet
        # styling is now cell-level: case_id + affected answer/reasoning cells
        # instead of whole-row review highlighting.
        red_review_rows = set(red_review_reasons)
        yellow_review_rows = set(yellow_review_reasons) - red_review_rows
        red_review_cells = _find_review_excel_cells(df, display_df, red_review_reasons)
        yellow_review_cells = _find_review_excel_cells(df, display_df, yellow_review_reasons)

        display_df.to_excel(out, index=False, sheet_name=sheet_name)

        if add_debug_sheet:
            _write_debug_summary_sheet(
                str(out),
                df,
                wrong_cells=wrong_cells,
                strange_reasons=yellow_review_reasons,
                red_review_reasons=red_review_reasons,
            )

        # Apply styling to the Excel file. Red highlights answer-key mismatches
        # and high-priority review rows; yellow highlights review-only rows.
        style_excel_sheet(
            str(out),
            sheet_name,
            wrong_cells=wrong_cells,
            strange_rows=yellow_review_rows,
            red_review_rows=red_review_rows,
            red_review_cells=red_review_cells,
            yellow_review_cells=yellow_review_cells,
        )
    elif fmt == "csv":
        display_df.to_csv(out, index=False)
    else:
        raise ValueError("Unsupported format: choose 'xlsx' or 'csv'")

    return out


if __name__ == "__main__":
    # Quick CLI for convenience.
    try:
        out_file = convert()
        print(f"Wrote spreadsheet: {out_file}")
    except Exception as e:
        print("Error:", e)
